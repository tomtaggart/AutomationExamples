#!/usr/bin/env python

import argparse
import logging
import logging.handlers
import datetime
import os
import smtplib
from email.mime.text import MIMEText
import pickle
import re

import openpyxl

import customer_automation_functions
import customer_account_variables
import customer_tfs_data_normalization


_TFS_SPREADSHEET = customer_account_variables.TFS_SPREADSHEET
_3rdPartyProgram_EMAIL = customer_account_variables.3rdPartyProgram_EMAIL
_DDTS_TACCASES_FILE = customer_account_variables.DDTS_TAC_CASE_PICKLE_JAR
_TFS_ADD_REQUEST_QUEUE_FILE = customer_account_variables.TFS_ADD_REQUEST_QUEUE_FILE

parser = argparse.ArgumentParser()
parser.add_argument("-d", "--DDTS", help="DDTS being submitted to TFS")
parser.add_argument("-e", "--email", default=None,
                    help="Email address where bug information will be sent.")
parser.add_argument("-l", "--loggingLevel",
                    choices=('DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'),
                    default='WARNING', help="Enable console logging level.")
parser.add_argument("-t", "--request_type", choices=['BUG', 'FEATURE'], \
                    help="DDTS is a bug or a feature request.")
args = parser.parse_args()

logger = logging.getLogger('TFS_ADD_REQUEST')
logger.setLevel(logging.DEBUG)
ch = logging.StreamHandler()
if args.loggingLevel == "WARNING":
    ch.setLevel(logging.WARNING)
elif args.loggingLevel == "DEBUG":
    ch.setLevel(logging.DEBUG)
elif args.loggingLevel == "INFO":
    ch.setLevel(logging.INFO)
elif args.loggingLevel == "ERROR":
    ch.setLevel(logging.ERROR)
else:
    ch.setLevel(logging.CRITICAL)
#fh = logging.FileHandler(customer_account_variables.TFS_LOGFILE)
fh = logging.handlers.RotatingFileHandler(customer_account_variables.TFS_LOGFILE, maxBytes=50000000, backupCount=5)
fh.setLevel(logging.DEBUG)
eh = logging.handlers.SMTPHandler('outbound.company.com', 'WARNING_DO_NOT_REPLY@company.com',
                                  ['user1@company.com'], 'Message from customer_TFS_add_request application')
eh.setLevel(logging.WARNING)
formatter = logging.Formatter('%(asctime)s [%(name)s] %(levelname)s: %(message)s')
fh.setFormatter(formatter)
ch.setFormatter(formatter)
eh.setFormatter(formatter)
logger.addHandler(fh)
logger.addHandler(ch)
logger.addHandler(eh)


if args.email != 'pythonScript':
    logger.info('\n\n\nAttempting to respond to TFS_ADDFEATURE_REQUEST for %s...' % args.email)
    ###Check if email sender is authorized to run application
    logger.info('Calling customer_automation_functions.check_sender()...')
    logger.debug('email=%s, variable_type=%s, loggingLevel=%s, variable_type=%s'
                  % (args.email, type(args.email), args.loggingLevel, type(args.loggingLevel)))
    if customer_automation_functions.check_sender(args.email.lower(), args.loggingLevel):
        logger.info('Creating DDTSList...')
        logger.debug('DDTSList=%s, variable_type=%s' % (args.DDTS, type(args.DDTS)))
        if args.DDTS:
            DDTSList = (args.DDTS).split(',')
            logger.info('DDTSList created.')
            logger.debug('DDTSList=%s, variable_type=%s' % (args.DDTS, type(args.DDTS)))
        else:
            logger.info('Application terminating due to no DDTS passed into application.')
            os.sys.exit(1)
        for (i, DDTS) in enumerate(DDTSList):
            DDTSList[i] = DDTS.strip()
        DDTS_TACCasesDict = False
    else:
        logger.info('Application terminating due to unauthorized email sender.')   
        logger.warning('unauthorized_email=%s has requested DDTS to be added to TFS database.' % args.email)
        os.sys.exit(1)    
else:
    #DDTS is coming from TAC case parsing.
    logger.info('\n\n\nAttempting to respond to TFS_ADDBUG_REQUEST from automated TAC parsing...')    
    DDTSList = []
    try:
        logger.info('Loading DDTS_TACCases.pkl...')
        f = open(_DDTS_TACCASES_FILE, 'rb')
        DDTS_TACCasesDict = pickle.load(f)
        f.close()
    except Exception as e:
        logger.info('Failed to open and load DDTS_TACCases.pkl pickle jar.')
        logger.debug(e)
    else:
        logger.info('Assigned DDTS_TACCasesDict.')
        logger.debug(DDTS_TACCasesDict)
    try:
        logger.info('Creating DDTSList...')
        for case in DDTS_TACCasesDict:
            DDTSList.append(DDTS_TACCasesDict[case].DDTS)
            logger.debug('Added DDTS %s to DDTSList.' % DDTS_TACCasesDict[case].DDTS)
    except Exception as e:
        logger.info('Failed to iterate through DDTS_TACCasesDict and create DDTSList.')
        logger.debug(e)
    else:
        logger.info('DDTSList created.')
        logger.debug(DDTSList)    
logger.info('Calling customer_automation_functions.validate_DDTS_ID()...')
logger.debug('DDTSList=%s, variable_type=%s' % (DDTSList, type(DDTSList)))
DDTSList = customer_automation_functions.validate_DDTS_ID(DDTSList, args.loggingLevel)
logger.debug(DDTSList)
if DDTSList == 0:
    logger.info('There are no DDTSs to process.  Terminating program.')
    os.sys.kill(1)


#Open TFS request queue.  The request queue is a pickle file tracking submissions the script
#has made in the past.  We do this to ensure we don't double submit DDTSs while the .xls
#spreadsheet is being synch'd.

try:
    logger.info("Opening tfs_add_request_queue.pkl file...")
    f = open(_TFS_ADD_REQUEST_QUEUE_FILE, 'rb')
    tfs_add_request_queue_dict = pickle.load(f)
    f.close()
except Exception as e:
    logger.info('Failed to open tfs_add_request_queue.pkl file.')
    logger.debug(e)
    os.sys.exit(1)
else:
    logger.info('tfs_add_request_queue_dict created')
    logger.debug(tfs_add_request_queue_dict)

#Open TFS spreadsheet and collect DDTS numbers and TFS IDs.  This will allow us
#to validate if a DDTS request we've made has been successful.

_start_openXLS_timestamp = datetime.datetime.now()
logger.info('Opening TFS spreadsheet...')
try:
    tfsWb = openpyxl.load_workbook(_TFS_SPREADSHEET)
except Exception as e:
    logger.info('Failed to open TFS spreadsheet.')
    logger.debug(e)
else:
    logger.info('TFS spreadsheet opened.')
tfsds = tfsWb.get_sheet_by_name('Defects')
tfsfs = tfsWb.get_sheet_by_name('Feature Requests')
logger.info('TFS spreadsheet opened.')
customer_automation_functions.runtime(_start_openXLS_timestamp, 'openXLS', 'spreadsheet', 1)

#Create successfully submitted TFS_ID/DDTS list for defects from TFS spreadsheet
DDTSsInTFSList = []
TFSIDsList = []
logger.info('Creating list of TFS IDs/DDTSs for defects that are in the TFS spreadsheet...')
for row in range(3, tfsds.max_row):
    try:
        currentTFSID = tfsds['A%i' % row].value
        currentDDTS = tfsds['B%i' % row].value
    except Exception as e:
        logger.info('Had issue reading cells from TFS spreadsheet.  Therefore, could not '
                    + 'create list of defect DDTSs that are in TFS from spreadsheet.  '
                    + 'Terminating program.')
        logger.debug(e)
        os.sys.exit(4)
    else:
        TFSIDsList.append(currentTFSID)
        DDTSsInTFSList.append(currentDDTS)
    
#Create successfully submitted TFS_ID/DDTS list for features from TFS spreadsheet
logger.info('Creating list of TFS IDs/DDTSs for features that are in the TFS spreadsheet...')
for row in range(3, tfsfs.max_row):
    try:
        currentTFSID = tfsfs['A%i' % row].value
        currentDDTS = tfsfs['B%i' % row].value
    except Exception as e:
        logger.info('Had issue reading cells from TFS spreadsheet.  Therefore, could not '
                    + 'create list of feature DDTSs that are in TFS from spreadsheet.  '
                    + 'Terminating program.')
        logger.debug(e)
        os.sys.exit(4)
    else:
        TFSIDsList.append(currentTFSID)
        DDTSsInTFSList.append(currentDDTS)
logger.info('DDTSsInTFSList and TFSIDsList have been created.')  
logger.debug(DDTSsInTFSList)
logger.debug(TFSIDsList)

#clean up tfs_add_request_queue by checking tfs_add_request_queue_list to see
#if DDTSs in request queue are in spreadsheet and have a TFS ID.
logger.info('Identifying DDTSs from tfs_add_request_queue_dict that are in TFS spreadsheet...')
DDTSsToBeRemoved = []
for DDTS in tfs_add_request_queue_dict:
    if DDTS in DDTSsInTFSList: 
        i = DDTSsInTFSList.index(DDTS)
        logger.info('DDTS %s is in DDTSsInTFSList at index %i with TFS ID %s' % (DDTS, i, TFSIDsList[i]))
        if TFSIDsList[i]:
            DDTSsToBeRemoved.append(DDTS)
            logger.info('DDTS %s is in TFS spreadsheet with a valid TFS ID and has been added to DDTSsToBeRemoved list.' % DDTS)
        else:
            logger.info('DDTS %s is in TFS spreadsheet without a valid TFS ID.  Either the spreadsheet needs to be ' % DDTS
                         + 'synchronized with TFS or TFS did not process the 3rdPartyProgram email request.')
    else:
        logger.info('DDTS %s is not in TFS spreadsheet.' % DDTS)
if len(DDTSsToBeRemoved) == 0:
    logger.info('There are no DDTS(s) to be removed from tfs_add_request_queue_dict.')
    _DDTS_removed = False
else:
    logger.info('Removing DDTSs that are in the TFS spreadsheet from tfs_add_request_queue_dict...')
    _DDTS_removed = True
    for DDTS in DDTSsToBeRemoved:
            #Send email to requester that DDTS is now in TFS
            try:
                success_msg = MIMEText('DDTS %s has been successfully added to the TFS system.' % DDTS)
            except Exception as e:
                    logger.warning('Successfully added to TFS message body could not be created '
                                   + 'for %s.' % tfs_add_request_queue_dict[DDTS]['requester'])
                    logger.debug(e)
            else: 
                success_msg['Subject'] = "DDTS %s Successfully added to TFS database." % DDTS
                success_msg['To'] = tfs_add_request_queue_dict[DDTS]['requester']
                try:
                    logger.info('Connecting to SMTP server...')
                    success_SMTP = smtplib.SMTP('outbound.company.com:25')
                except Exception as e:
                    logger.warning('Connection to Company SMTP server failed for user=%s ' % tfs_add_request_queue_dict[DDTS]['requester']
                                   + 'for DDTS successfully added email.')
                    logger.debug(e)
                else:
                    logger.info('Connection to SMTP server complete.')
                    try:
                        logger.info('Sending "successfully added to TFS" email to Company SMTP server.')
                        #Added below if statement for cases where a DDTS is added via the pythonScript and therefore there is no email
                        #address to reply to.
                        if tfs_add_request_queue_dict[DDTS]['requester'] != 'pythonScript':    
                            success_SMTP.sendmail('user1@company.com', tfs_add_request_queue_dict[DDTS]['requester'], success_msg.as_string())
                    except Exception as e:
                        logger.info('"Successfully added to TFS" message could not be sent to outbound.company.com:25.  Terminating application.')
                        logger.debug(e)
                        logger.warning('Outbound SMTP connection, outbound.company.com:25, has failed for DDTS successfully added email ' 
                                       + 'to %s for %s' % tfs_add_request_queue_dict[DDTS]['requester'], DDTS)
                    else:
                        logger.info('Successfully added to TFS email sent to %s.' % tfs_add_request_queue_dict[DDTS]['requester'])
                    if success_SMTP:
                        success_SMTP.quit() 
            tfs_add_request_queue_dict.pop(DDTS)
            logger.info('DDTS %s removed from tfs_add_request_queue_dict'% DDTS)

for DDTS in DDTSList:
    if DDTS in DDTSsInTFSList:
        logger.info('DDTS %s is already in the TFS system.' % DDTS)
        logger.info('Building DDTS already in the TFS system email...')
        if args.email != 'pythonScript':
            #Send email to requester that DDTS is now in TFS
            try:
                inTFS_msg = MIMEText('DDTS %s is already in the TFS system.' % DDTS)
            except Exception as e:
                    logger.warning('Already in TFS message body could not be created '
                                   + 'for %s.' % args.email)
                    logger.debug(e)
            else: 
                inTFS_msg['Subject'] = "TFS_ADD_%s_RESPONSE=%s" % (args.request_type.upper(), DDTS)
                inTFS_msg['To'] = args.email
                try:
                    logger.info('Connecting to SMTP server...')
                    inTFS_SMTP = smtplib.SMTP('outbound.company.com:25')
                except Exception as e:
                    logger.warning('Connection to Company SMTP server failed for user=%s' % args.email)
                    logger.debug(e)
                else:
                    try:
                        logger.info('Sending successfully already in TFS email to Company SMTP server.')
                        inTFS_SMTP.sendmail('user1@company.com', args.email, inTFS_msg.as_string())
                    except Exception as e:
                        logger.info('Already in TFS message could not be sent to outbound.company.com:25.  Terminating application.')
                        logger.debug(e)
                        logger.warning('Outbound SMTP connection, outbound.company.com:25, has failed for already-in TFS email. ' 
                                       + 'to %s for %s' % (args.email, DDTS))
                    else:
                        logger.info('Successfully sent already in TFS email to %s.' % args.email)
                    if inTFS_SMTP:
                        inTFS_SMTP.quit() 
        #For future, if tac case parsing sends in DDTS, reply to that TAC case with the TFS ID.  Need to do this ONLY once.
        #    customer_automation_functions.TFS_ID_attach(DDTS, TFS_ID, DDTS_TACCasesDict[DDTS]['caseNumber'], args.loggingLevel)
    elif DDTS not in DDTSsInTFSList and tfs_add_request_queue_dict.has_key(DDTS):
        logger.info('DDTS %s of type, %s, was already added to tfs_add_request_queue on %s by %s'
                    % (DDTS, tfs_add_request_queue_dict[DDTS]['type'],
                       tfs_add_request_queue_dict[DDTS]['request_date'],
                       tfs_add_request_queue_dict[DDTS]['requester']
                      )
                    )
        if args.email != 'pythonScript':
            logger.info('Building DDTS already in Queue Email...')
            try:
                inQueue_msg = MIMEText('DDTS %s is already in the TFS_ADD_QUEUE.  It was added ' % DDTS
                                       + 'by %s on %s' % (tfs_add_request_queue_dict[DDTS]['requester'],
                                                          tfs_add_request_queue_dict[DDTS]['request_date']
                                                          )
                                       )
            except Exception as e:
                    logger.warning('Already in Queue message body could not be created '
                                   + 'for %s.' % args.email)
                    logger.debug(e)
            else: 
                inQueue_msg['Subject'] = "TFS_ADD_%s_RESPONSE=%s" % (args.request_type.upper(), DDTS)
                inQueue_msg['To'] = args.email
                try:
                    logger.info('Connecting to SMTP server...')
                    inQueue_SMTP = smtplib.SMTP('outbound.company.com:25')
                except Exception as e:
                    logger.warning('Connection to Company SMTP server failed for user=%s' % args.email)
                    logger.debug(e)
                else:
                    try:
                        logger.info('Sending already in TFS email to Company SMTP server.')
                        inQueue_SMTP.sendmail('user1@company.com', args.email, inQueue_msg.as_string())
                    except Exception as e:
                        logger.info('Already-in TFS message could not be sent to outbound.company.com:25.')
                        logger.debug(e)
                        logger.warning('Outbound SMTP connection, outbound.company.com:25, has failed for already '
                                       + 'in TFS_ADD_QUEUE added email to %s for %s' % (args.email, DDTS))
                    else:
                        logger.info('Successfully sent already in TFS email to %s.' % args.email)
                    if inQueue_SMTP:
                        inQueue_SMTP.quit() 
    else: #DDTS not in DDTSsInTFSList and not tfs_add_request_queue_dict.has_key(DDTS)
        #start to build 3rdPartyProgram message body.  Make calls to customer_tfs_data_normalization.py.
        try:
            DDTSDetails = customer_automation_functions.automation system_fasterbug(DDTS, args.loggingLevel)
            '''
            #Removing the call to customer_automation_functions.automation system_fasterbug_releasenote_raw as it is failing for newer bugs.
            #this needs to be rewritten with teh bug tracking program CLI as AUTOMATION SYSTEM call is unreliable.
            DDTSDetails['ReleaseNote'] = customer_automation_functions.automation system_fasterbug_releasenote_raw(DDTS, args.loggingLevel)
            #if $$IGNORE in release-note return a generic description as this may be PSIRT or sensitive.
            if '$$IGNORE' in DDTSDetails['ReleaseNote'] or '$$PREFCS' in DDTSDetails['ReleaseNote']:
                DDTSDetails['ReleaseNote'] = 'Company Team Validating...'
            '''
            DDTSDetails['ReleaseNote'] = 'https://bst.cloudapps.company.com/bugsearch/bug/%s' % DDTS
            logger.info('Communication to AUTOMATION SYSTEM complete.')
            logger.debug(DDTSDetails)
            #if release notes note other DDTSs this DDTS is most likely an umbrella DDTS and we need to add the additional
            #DDTSs to the DDTSList.
            if 'csc' in DDTSDetails['ReleaseNote'].lower():
                umbrellaDDTSList = []
                for m in re.finditer('csc', DDTSDetails['ReleaseNote'].lower()):
                    umbrellaDDTSList.append(DDTSDetails['ReleaseNote'][m.start(): m.start() + 10])
                    umbrellaDDTSList = customer_automation_functions.validate_DDTS_ID(umbrellaDDTSList, args.loggingLevel)
                logger.info('Umbrella DDTS list created.')
                logger.debug(umbrellaDDTSList)
                DDTSList += umbrellaDDTSList
                logger.debug('Adding umbrella DDTS list to DDTSList...')
                logger.debug(DDTSList)
            
            if args.request_type == 'BUG': #feature won't have these.  Also, DDTS_TACCaseDict came from Python script not email.
                _caseNumber_set = False
                _hostname_set = False
                DDTSDetails['hostname'] =  'TBD'
                DDTSDetails['caseNumber'] = 'TBD'
                if DDTS_TACCasesDict:
                    for case in DDTS_TACCasesDict:
                        if DDTS_TACCasesDict[case].DDTS == DDTS:
                            if DDTS_TACCasesDict[case].hostname:
                                DDTSDetails['hostname'] = customer_tfs_data_normalization.node_name(DDTS_TACCasesDict[case].hostname)
                                _hostname_set = True
                            DDTSDetails['caseNumber'] = DDTS_TACCasesDict[case].caseNumber
                            _caseNumber_set = True  
                #if not _caseNumber_set:
                 #   DDTSDetails['hostname'] =  'TBD'
                #if not _hostname_set:
                 #   DDTSDetails['caseNumber'] = 'TBD'

            DDTSDetails['review_required'] = customer_tfs_data_normalization.review_required('initialAdd')
            DDTSDetails['review_required_comments'] = customer_tfs_data_normalization.review_required_comments('initialAdd')
            logger.info('Calling customer_tfs_data_normalization.product_stats()...')
            logger.debug('hostname=%s' % DDTSDetails['hostname'])
            stats = customer_tfs_data_normalization.product_stats(DDTSDetails['product'], DDTSDetails['hostname'])
            logger.debug(stats)
            DDTSDetails['product_group'] = stats[0]
            DDTSDetails['product_model'] = stats[1]
            DDTSDetails['os'] = stats[2]
            DDTSDetails['customer_POC'] = customer_tfs_data_normalization.customer_POC(DDTSDetails['product'], DDTSDetails['hostname'])
            DDTSDetails['supplier_POC'] =  customer_tfs_data_normalization.supplier_POC(DDTSDetails['product'], DDTSDetails['hostname'])

        except Exception as e:
            logger.info('Could not collect bug details and craft email for 3rdPartyProgram system.')
            logger.exception(e)
        else:
            logger.info('Message body for 3rdPartyProgram system created.')
            logging.debug(DDTSDetails)
        
        logger.info('Building email to 3rdPartyProgram system...')
        if args.request_type == 'BUG':
            try:
                M2B_msg = MIMEText(DDTSDetails['ReleaseNote'] + "\n\n"
                                   + '###Supplier Defect ID:' + DDTSDetails['DDTSNumber'] + "\n"
                                   + '###Title:' + DDTSDetails['DDTSTitle'] + "\n"
                                   + '###TAC Number(s):' + DDTSDetails['caseNumber'] + "\n"
                                   + '###Supplier:Company' + "\n"
                                   + '###Environment Found:Confirming' + "\n"
                                   + '###Found in Software Version:' + DDTSDetails['found_in_sw'] + "\n"
                                   #+ '###Area Path:GNS-Lab\Suppliers\Company' + "\n"
                                   #+ '###Node Name:' + DDTSDetails['hostname'] + "\n"
                                   + '###Review Required:' + DDTSDetails['review_required'] + "\n"
                                   + '###Review Required Comments:' + DDTSDetails['review_required_comments'] + "\n"
                                   + '###Product Group:' + DDTSDetails['product_group'] + "\n"
                                   + '###Model:' + DDTSDetails['product_model'] + "\n"
                                   + '###OS:' + DDTSDetails['os'] + "\n"
                                   + '###MS Points of Contact:' + DDTSDetails['customer_POC'] + "\n"
                                   + '###Supplier POC:' + DDTSDetails['supplier_POC'] + "\n"
                                   + '###Integrated Release:' + DDTSDetails['integrated'] + "\n"
                                   #+ '###Workarounds:' + DDTSDetails['ReleaseNote'][(DDTSDetails['ReleaseNote'].find('orkaround')) + 10:] + "\n"
                                   + '###Workarounds:See Description\n'
                                   #+ '###Description:' + DDTSDetails['ReleaseNote'][0: (DDTSDetails['ReleaseNote'].find('orkaround') - 1)] + "\n"
                                   + '\n'
                                   #+ DDTSDetails['ReleaseNote'] # moved this to top of line to seeif 3rdPartyProgram puts in the description field.
                                  )
            except Exception as e:
                    logger.warning('TFS message body for DDTS %s bug could not be created for 3rdPartyProgram system.' % DDTS)
                    logger.debug(e)
            else:
                logger.info('3rdPartyProgram message body for defect %s created.' % DDTS)
        else:
            try:
                #content for feature email body.
                M2B_msg = MIMEText('###Supplier Defect ID:' + DDTSDetails['DDTSNumber'] + "\n"
                                   + '###Title:' + DDTSDetails['DDTSTitle'] + "\n"
                                   + '###Area Path:GNS-Lab\Suppliers\Company' + "\n"
                                   + '###Supplier:Company' + "\n"
                                   + '\n\n'
                                   + DDTSDetails['ReleaseNote'] 
                                  )
            except Exception as e:
                    logger.warning('TFS message body for DDTS %s feature could not be created for 3rdPartyProgram system.' % DDTS)
                    logger.debug(e)
            else:
                logger.info('3rdPartyProgram message body for feature %s created.' % DDTS)
        if args.request_type == 'BUG':
            M2B_msg['Subject'] = "[company defect]"
        else:
            M2B_msg['Subject'] = "[company feature]"
        M2B_msg['To'] = _3rdPartyProgram_EMAIL
        try:
            logger.info('Connecting to SMTP server...')
            M2B_SMTP = smtplib.SMTP('outbound.company.com:25')
        except Exception as e:
            logger.warning('Connection to Company SMTP server failed for user=%s' % args.email)
            logger.debug(e)
        else:
            try:
                logger.info('Sending 3rdPartyProgram email to Company SMTP server.')
                M2B_SMTP.sendmail('user1@company.com',_3rdPartyProgram_EMAIL, M2B_msg.as_string())
            except Exception as e:
                logger.info('3rdPartyProgram message could not be sent to outbound.company.com:25.  Terminating application.')
                logger.debug(e)
                logger.warning('Outbound SMTP connection has failed for outbound.company.com:25 for 3rdPartyProgram email ' 
                               + 'to %s for %s' % (_3rdPartyProgram_EMAIL, DDTS))
            else:
                logger.info('3rdPartyProgram email sent to %s.' % _3rdPartyProgram_EMAIL)
            if M2B_SMTP:
                M2B_SMTP.quit()    
        
        logger.info('Calling customer_automation_functions.purge_DDTS_email(%s)' % DDTS)
        customer_automation_functions.purge_DDTS_email(DDTS, args.loggingLevel)
        
        #add DDTS to tfs_request_queue_dict
        tfs_add_request_queue_dict[DDTS] = {'type': args.request_type, 'request_date': datetime.datetime.now(),
                                            'ID': DDTS, 'requester': args.email, 'caseNumber': DDTSDetails['caseNumber'],
                                            'hostname': DDTSDetails['hostname']}

        #send email to requester if one exists.
        if args.email != 'pythonScript':
            logger.info('Building TFS_ADD_RESPONSE email to %s...' % args.email)
            try:
                TFS_msg = MIMEText('Hello,\n\n'
                                   + 'You requested via email that DDTS %s be submitted to ' % DDTS
                                   + 'the TFS system.  This email is a reply to let you know that '
                                   + 'the details for the DDTS have been collected internally at '
                                   + 'Company and an email has been sent to the 3rdPartyProgram application for '
                                   + 'final addition to the TFS system.  If you have any questions or '
                                   + 'need further clarification please respond to this email describing '
                                   + 'what you need and we will be in touch with you ASAP.  Thank You.')
            except Exception as e:
                    logger.warning('TFS_ADD_Response message body could not be created for %s.' % args.email)
                    logger.debug(e)
            else: 
                TFS_msg['Subject'] = "TFS_ADD_%s_RESPONSE=%s" % (args.request_type.upper(), DDTS)
                TFS_msg['To'] = args.email
                try:
                    logger.info('Connecting to SMTP server...')
                    TFS_SMTP = smtplib.SMTP('outbound.company.com:25')
                except Exception as e:
                    logger.warning('Connection to Company SMTP server failed for user=%s' % args.email.upper())
                    logger.debug(e)
                else:
                    try:
                        logger.info('Sending TFS_ADD_RESPONSE email to Company SMTP server.')
                        TFS_SMTP.sendmail('user1@company.com',args.email, TFS_msg.as_string())
                    except Exception as e:
                        logger.info('TFS_RESPONSE message could not be sent to outbound.company.com:25.  Terminating application.')
                        logger.debug(e)
                        logger.warning('Outbound SMTP connection has failed for outbound.company.com:25 for TFS_ADD_RESPONSE email ' 
                                       + 'to %s for %s.' % (_3rdPartyProgram_EMAIL, DDTS))
                    else:
                        logger.info('TFS_ADD_RESPONES email sent to %s.' % args.email)
                        if TFS_SMTP:
                            TFS_SMTP.quit()


        #Pickle new tfs_request_queue_dict
        _deletedQueueFile = False
        try:
            logger.info('Checking for existing TFS_REQUEST_QUEUE pickle jar and erasing if found.')
            if os.path.exists(_TFS_ADD_REQUEST_QUEUE_FILE):
                os.remove(_TFS_ADD_REQUEST_QUEUE_FILE)
                _deletedQueueFile = True
        except Exception as e:
            logger.info('Failed to check for or remove %s.' % _TFS_ADD_REQUEST_QUEUE_FILE)
            logger.debug(e)
        else:
            if _deletedQueueFile:
                logger.debug('Deleted %s.' % _TFS_ADD_REQUEST_QUEUE_FILE)
        
        try:
            logger.info('Attempting to pickle tfs_add_request_queue_dict...')
            f = open(_TFS_ADD_REQUEST_QUEUE_FILE, 'wb')
            pickle.dump(tfs_add_request_queue_dict, f)
            f.close()
        except Exception as e:
            logger.info('Failed to pickle %s.' % _TFS_ADD_REQUEST_QUEUE_FILE)
            logger.debug(e)
        else:
            logger.info('%i case(s) added to tfs_add_request_queue.pkl' % len(tfs_add_request_queue_dict))
            logger.info('Pickled %s.' % _TFS_ADD_REQUEST_QUEUE_FILE)
            

#clean up and delete DDTS_TACCaseDict.pkl from queue directory.
if args.email == 'pythonScript':
    _deletedDDTSFile = False
    try:
        logger.info('Checking for existing DDTS_TACCasesDict pickle jar and erasing if found.')
        if os.path.exists(_DDTS_TACCASES_FILE):
            os.remove(_DDTS_TACCASES_FILE)
            _deletedDDTSFile = True
    except Exception as e:
        logger.info('Failed to check for or remove %s.', _DDTS_TACCASES_FILE, exc_info=True)
    else:
        if _deletedDDTSFile:
            logger.debug('Deleted %s.' % _DDTS_TACCASES_FILE)